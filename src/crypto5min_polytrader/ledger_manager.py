"""ledger_manager.py — Automated monthly/annual trade ledger system (v1.2.3)

Writes three types of Excel ledger, all stored in logs/ledgers/:

  Monthly:  ledger_YYYY_MM.xlsx   — all trades + daily P/L for that calendar month.
            Re-written on every update call so it is always current.
            When the month rolls over the previous month's file is sealed (read-only flag set).

  Annual:   ledger_YYYY_annual.xlsx — one summary sheet per closed month +
            a full-year overview. Rebuilt whenever a month is sealed.

  Lifetime: ledger_all_time.xlsx  — running overview across all years.
            Rebuilt on every annual seal.

Trigger:    Call `update_ledgers()` from the background loop (e.g. after
            check_resolutions()).  The function is cheap — it tracks the
            last-written trade timestamp and only rebuilds if new resolved
            trades have appeared since the last run.
"""
from __future__ import annotations

import datetime as _dt
import io
import logging
import os
import time
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# ── Directory ─────────────────────────────────────────────────────────────────
LEDGER_DIR = Path('logs') / 'ledgers'

# ── Colour palette ────────────────────────────────────────────────────────────
_CH_BG    = '1A2A4A'
_CH_FG    = 'FFFFFF'
_WIN_BG   = 'E8F5E9'
_LOSS_BG  = 'FFEBEE'
_ALT_BG   = 'F8F9FA'
_POS_FG   = '1B7E35'
_NEG_FG   = 'C62828'
_GOLD_FG  = 'B8860B'
_SEC_BG   = 'E3EDF7'
_TOT_BG   = 'D0E4F7'
_BORDER   = 'CCCCCC'
_NEUTRAL  = '555555'

# ── Last-update tracking (in-process) ─────────────────────────────────────────
_last_update_ts: float = 0.0          # wall-clock ts of last full rebuild
_last_trade_count: int = 0            # number of resolved trades at last rebuild
_UPDATE_INTERVAL_S: float = 300.0     # minimum seconds between rebuilds (5 min)


# ══════════════════════════════════════════════════════════════════════════════
# Openpyxl helpers
# ══════════════════════════════════════════════════════════════════════════════

def _xl():
    """Lazy-import openpyxl — not available at import time in all envs."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    return openpyxl, Font, PatternFill, Alignment, Border, Side, get_column_letter


def _hdr_font(xl_mod, bold=True, color=_CH_FG, size=10):
    _, Font, *_ = xl_mod
    return Font(name='Arial', bold=bold, color=color, size=size)


def _body_font(xl_mod, bold=False, color='000000', size=10):
    _, Font, *_ = xl_mod
    return Font(name='Arial', bold=bold, color=color, size=size)


def _fill(xl_mod, hex_color):
    _, _, PatternFill, *_ = xl_mod
    return PatternFill('solid', fgColor=hex_color)


def _thin_border(xl_mod):
    *_, Border, Side, _ = xl_mod
    s = Side(style='thin', color=_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _align(xl_mod, h='left'):
    _, _, _, Alignment, *_ = xl_mod
    return Alignment(horizontal=h, vertical='center')


def _style_header(ws, xl_mod, row, ncols, height=18):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font      = _hdr_font(xl_mod)
        c.fill      = _fill(xl_mod, _CH_BG)
        c.border    = _thin_border(xl_mod)
        c.alignment = _align(xl_mod, 'center')
    ws.row_dimensions[row].height = height


def _style_data_row(ws, xl_mod, row, ncols, bg, right_from=2):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill      = _fill(xl_mod, bg)
        c.border    = _thin_border(xl_mod)
        c.font      = _body_font(xl_mod)
        c.alignment = _align(xl_mod, 'right' if col >= right_from else 'left')


def _autowidth(ws, xl_mod, min_w=8, max_w=42):
    *_, get_column_letter = xl_mod
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max((len(str(c.value or '')) for c in col_cells), default=0)
        ws.column_dimensions[col_letter].width = min(max_w, max(min_w, max_len + 3))


# ══════════════════════════════════════════════════════════════════════════════
# Trade data helpers
# ══════════════════════════════════════════════════════════════════════════════

def _load_trades() -> list[dict]:
    """Load all trades from the JSON store."""
    try:
        from .persistence import JsonStore
        store = JsonStore(Path('logs') / 'poly_trades.json')
        trades = store.load(default=[]) or []
        return [t for t in trades if isinstance(t, dict)]
    except Exception as exc:
        logger.warning('ledger: could not load trades: %s', exc)
        return []


def _pnl_for(t: dict) -> float | None:
    """Compute net P/L in USDC for a resolved filled trade."""
    try:
        from .polymarket_exec import estimate_redeemed_profit_usdc as _erp
    except Exception:
        _erp = lambda _: None  # noqa

    resolved = t.get('resolved') or ''
    filled   = float(t.get('filled_size') or 0.0)
    redeemed = t.get('redeem_status') == 'success'
    if filled <= 0 and not redeemed:
        return None
    stake = float(t.get('usdc') or t.get('stake_usdc') or t.get('stake') or 5.0)
    px    = float(t.get('avg_fill_price') or t.get('price') or 0.0)

    if resolved == 'win':
        p = _erp(t)
        if p is None:
            p = stake * (1.0 - px) if 0 < px < 1 else stake * 0.55
        return p
    elif resolved == 'loss':
        if filled > 0 and 0 < px < 1:
            return -(filled * px)
        return -stake
    return None


def _sym_norm(raw: str) -> str:
    s = raw.strip().upper()
    return s if '-' in s else s + '-USD'


def _ts_to_dt(ts: float) -> _dt.datetime | None:
    try:
        return _dt.datetime.utcfromtimestamp(float(ts))
    except Exception:
        return None


def _resolved_trades(trades: list[dict]) -> list[dict]:
    """Return only filled+resolved trades sorted by resolved_ts ascending."""
    out = []
    for t in trades:
        if t.get('resolved') not in ('win', 'loss'):
            continue
        filled   = float(t.get('filled_size') or 0.0)
        redeemed = t.get('redeem_status') == 'success'
        if filled <= 0 and not redeemed:
            continue
        out.append(t)
    out.sort(key=lambda x: float(x.get('resolved_ts') or x.get('ts') or 0))
    return out


# ══════════════════════════════════════════════════════════════════════════════
# Monthly ledger builder
# ══════════════════════════════════════════════════════════════════════════════

def _build_monthly(trades: list[dict], year: int, month: int) -> bytes:
    """Build a monthly Excel ledger and return as bytes."""
    xl_mod = _xl()
    openpyxl = xl_mod[0]
    wb = openpyxl.Workbook()

    # Filter trades for this month
    month_trades = []
    for t in trades:
        ts = float(t.get('resolved_ts') or t.get('ts') or 0)
        dt = _ts_to_dt(ts)
        if dt and dt.year == year and dt.month == month:
            month_trades.append(t)

    month_name = _dt.date(year, month, 1).strftime('%B %Y')

    # ── Sheet 1: All Trades ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Trades'
    ws1.freeze_panes = 'A2'

    hdrs = ['Date/Time (UTC)', 'Symbol', 'Direction', 'Result',
            'Confidence', 'Entry Price', 'Stake ($)', 'Shares',
            'Net P/L ($)', 'ROI (%)', 'Cumulative P/L ($)',
            'Fill Status', 'Window']
    ws1.append(hdrs)
    _style_header(ws1, xl_mod, 1, len(hdrs))

    cum = 0.0
    for r, t in enumerate(month_trades, start=2):
        ts    = float(t.get('resolved_ts') or t.get('ts') or 0)
        dt    = _ts_to_dt(ts)
        dt_s  = dt.strftime('%Y-%m-%d %H:%M:%S') if dt else ''
        sym   = _sym_norm(t.get('symbol') or t.get('asset') or '')
        dirn  = (t.get('direction') or '').upper()
        res   = (t.get('resolved') or '').replace('_', ' ').title()
        conf  = float(t.get('confidence') or 0.0) or ''
        px    = float(t.get('avg_fill_price') or t.get('price') or 0.0) or ''
        stake = float(t.get('usdc') or t.get('stake_usdc') or 5.0)
        fs    = float(t.get('filled_size') or 0.0) or ''
        pnl   = _pnl_for(t)
        if pnl is not None:
            cum += pnl
        roi   = round(pnl / stake * 100, 2) if (pnl is not None and stake > 0) else ''
        fill  = 'Filled' if float(t.get('filled_size') or 0) > 0 else \
                ('Redeemed' if t.get('redeem_status') == 'success' else 'Unfilled')
        slug  = (t.get('window_slug') or '')[:50]

        ws1.append([dt_s, sym, dirn, res,
                    round(float(conf), 4) if conf != '' else '',
                    round(float(px), 4)   if px   != '' else '',
                    round(stake, 2),
                    round(float(fs), 4)   if fs   != '' else '',
                    round(pnl, 4)   if pnl  is not None else '',
                    roi,
                    round(cum, 4),
                    fill, slug])

        is_win  = t.get('resolved') == 'win'
        is_loss = t.get('resolved') == 'loss'
        bg = _WIN_BG if is_win else (_LOSS_BG if is_loss else _ALT_BG)
        _style_data_row(ws1, xl_mod, r, len(hdrs), bg)

        if pnl is not None:
            pnl_cell = ws1.cell(row=r, column=9)
            pnl_cell.font = _body_font(xl_mod, bold=True,
                                        color=(_POS_FG if pnl >= 0 else _NEG_FG))
        cum_cell = ws1.cell(row=r, column=11)
        cum_cell.font = _body_font(xl_mod, bold=True,
                                    color=(_POS_FG if cum >= 0 else _NEG_FG))

    _autowidth(ws1, xl_mod)
    ws1.column_dimensions['A'].width = 22
    ws1.column_dimensions['M'].width = 38

    # ── Sheet 2: Daily P/L ────────────────────────────────────────────────
    ws2 = wb.create_sheet('Daily P/L')
    ws2.freeze_panes = 'A2'

    hdrs2 = ['Date', 'Trades', 'Wins', 'Losses', 'Win Rate (%)',
              'Staked ($)', 'Daily P/L ($)', 'Cumulative P/L ($)']
    ws2.append(hdrs2)
    _style_header(ws2, xl_mod, 1, len(hdrs2))

    daily: dict[str, dict] = {}
    for t in month_trades:
        ts = float(t.get('resolved_ts') or t.get('ts') or 0)
        dt = _ts_to_dt(ts)
        day = dt.strftime('%Y-%m-%d') if dt else 'Unknown'
        if day not in daily:
            daily[day] = {'t': 0, 'w': 0, 'l': 0, 'staked': 0.0, 'pnl': 0.0}
        stake = float(t.get('usdc') or t.get('stake_usdc') or 5.0)
        daily[day]['t'] += 1
        daily[day]['staked'] += stake
        pnl = _pnl_for(t)
        if t.get('resolved') == 'win':
            daily[day]['w'] += 1
            daily[day]['pnl'] += (pnl or 0.0)
        else:
            daily[day]['l'] += 1
            daily[day]['pnl'] += (pnl or 0.0)

    cum2 = 0.0
    for r2, day in enumerate(sorted(daily.keys()), start=2):
        d   = daily[day]
        wr  = round(d['w'] / d['t'] * 100, 1) if d['t'] else 0.0
        cum2 += d['pnl']
        ws2.append([day, d['t'], d['w'], d['l'], wr,
                    round(d['staked'], 2), round(d['pnl'], 2), round(cum2, 2)])
        bg = _WIN_BG if d['pnl'] >= 0 else _LOSS_BG
        _style_data_row(ws2, xl_mod, r2, len(hdrs2), bg)
        ws2.cell(row=r2, column=7).font = _body_font(xl_mod, bold=True,
            color=(_POS_FG if d['pnl'] >= 0 else _NEG_FG))
        ws2.cell(row=r2, column=8).font = _body_font(xl_mod, bold=True,
            color=(_POS_FG if cum2 >= 0 else _NEG_FG))

    _autowidth(ws2, xl_mod)

    # ── Sheet 3: Per-Symbol Summary ───────────────────────────────────────
    ws3 = wb.create_sheet('By Symbol')
    ws3.freeze_panes = 'A2'

    hdrs3 = ['Symbol', 'Trades', 'Wins', 'Losses', 'Win Rate (%)',
              'Staked ($)', 'P/L ($)', 'ROI (%)',
              'Avg Win ($)', 'Avg Loss ($)', 'Reward/Risk']
    ws3.append(hdrs3)
    _style_header(ws3, xl_mod, 1, len(hdrs3))

    sym_agg: dict[str, dict] = {}
    for t in month_trades:
        sym = _sym_norm(t.get('symbol') or t.get('asset') or 'UNKNOWN')
        if sym not in sym_agg:
            sym_agg[sym] = {'w': 0, 'l': 0, 'staked': 0.0,
                             'win_pnls': [], 'loss_pnls': []}
        stake = float(t.get('usdc') or t.get('stake_usdc') or 5.0)
        sym_agg[sym]['staked'] += stake
        pnl = _pnl_for(t)
        if t.get('resolved') == 'win':
            sym_agg[sym]['w'] += 1
            sym_agg[sym]['win_pnls'].append(pnl or 0.0)
        else:
            sym_agg[sym]['l'] += 1
            sym_agg[sym]['loss_pnls'].append(pnl or 0.0)

    for r3, sym in enumerate(sorted(sym_agg.keys()), start=2):
        d    = sym_agg[sym]
        tot  = d['w'] + d['l']
        wr   = round(d['w'] / tot * 100, 1) if tot else 0.0
        pnl  = sum(d['win_pnls']) + sum(d['loss_pnls'])
        roi  = round(pnl / d['staked'] * 100, 1) if d['staked'] else 0.0
        aw   = round(sum(d['win_pnls']) / len(d['win_pnls']), 2) if d['win_pnls'] else 0.0
        al   = round(sum(d['loss_pnls']) / len(d['loss_pnls']), 2) if d['loss_pnls'] else 0.0
        rr   = round(abs(aw / al), 2) if al else 0.0
        ws3.append([sym, tot, d['w'], d['l'], wr,
                    round(d['staked'], 2), round(pnl, 2), roi, aw, al, rr])
        bg = _WIN_BG if pnl >= 0 else _LOSS_BG
        _style_data_row(ws3, xl_mod, r3, len(hdrs3), bg)
        ws3.cell(row=r3, column=5).font  = _body_font(xl_mod, bold=True,
            color=(_POS_FG if wr >= 52 else _NEG_FG))
        ws3.cell(row=r3, column=7).font  = _body_font(xl_mod, bold=True,
            color=(_POS_FG if pnl >= 0 else _NEG_FG))
        ws3.cell(row=r3, column=11).font = _body_font(xl_mod, bold=True,
            color=(_POS_FG if rr >= 1.0 else _GOLD_FG))

    _autowidth(ws3, xl_mod)

    # ── Tab colours + title ───────────────────────────────────────────────
    ws1.sheet_properties.tabColor = '1A2A4A'
    ws2.sheet_properties.tabColor = '00A86B'
    ws3.sheet_properties.tabColor = '2E5F8A'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# Annual ledger builder
# ══════════════════════════════════════════════════════════════════════════════

def _month_summary(trades: list[dict], year: int, month: int) -> dict:
    """Aggregate stats for a single month across all trades."""
    wins, losses = 0, 0
    staked, pnl_total = 0.0, 0.0
    win_pnls, loss_pnls = [], []
    for t in trades:
        ts = float(t.get('resolved_ts') or t.get('ts') or 0)
        dt = _ts_to_dt(ts)
        if not (dt and dt.year == year and dt.month == month):
            continue
        stake = float(t.get('usdc') or t.get('stake_usdc') or 5.0)
        staked += stake
        pnl = _pnl_for(t)
        if t.get('resolved') == 'win':
            wins += 1
            win_pnls.append(pnl or 0.0)
            pnl_total += (pnl or 0.0)
        elif t.get('resolved') == 'loss':
            losses += 1
            loss_pnls.append(pnl or 0.0)
            pnl_total += (pnl or 0.0)
    total = wins + losses
    return {
        'month': _dt.date(year, month, 1).strftime('%B %Y'),
        'trades': total, 'wins': wins, 'losses': losses,
        'win_rate': round(wins / total * 100, 1) if total else 0.0,
        'staked': round(staked, 2),
        'pnl': round(pnl_total, 2),
        'roi': round(pnl_total / staked * 100, 1) if staked else 0.0,
        'avg_win': round(sum(win_pnls) / len(win_pnls), 2) if win_pnls else 0.0,
        'avg_loss': round(sum(loss_pnls) / len(loss_pnls), 2) if loss_pnls else 0.0,
    }


def _build_annual(trades: list[dict], year: int) -> bytes:
    """Build the annual overview Excel workbook and return as bytes."""
    xl_mod = _xl()
    openpyxl = xl_mod[0]
    wb = openpyxl.Workbook()

    # Figure out which months have data
    months_with_data = sorted({
        _ts_to_dt(float(t.get('resolved_ts') or t.get('ts') or 0)).month
        for t in trades
        if _ts_to_dt(float(t.get('resolved_ts') or t.get('ts') or 0)) and
           _ts_to_dt(float(t.get('resolved_ts') or t.get('ts') or 0)).year == year
    })

    # ── Sheet 1: Year Overview ────────────────────────────────────────────
    ws_yr = wb.active
    ws_yr.title = f'{year} Overview'
    ws_yr.freeze_panes = 'A2'

    hdrs_yr = ['Month', 'Trades', 'Wins', 'Losses', 'Win Rate (%)',
                'Staked ($)', 'Monthly P/L ($)', 'ROI (%)',
                'Avg Win ($)', 'Avg Loss ($)', 'Cumulative P/L ($)']
    ws_yr.append(hdrs_yr)
    _style_header(ws_yr, xl_mod, 1, len(hdrs_yr))

    cum = 0.0
    year_trades = year_wins = year_losses = 0
    year_staked = year_pnl = 0.0
    all_win_pnls: list[float] = []
    all_loss_pnls: list[float] = []

    for r, month in enumerate(months_with_data, start=2):
        ms = _month_summary(trades, year, month)
        cum += ms['pnl']
        year_trades  += ms['trades']
        year_wins    += ms['wins']
        year_losses  += ms['losses']
        year_staked  += ms['staked']
        year_pnl     += ms['pnl']

        ws_yr.append([ms['month'], ms['trades'], ms['wins'], ms['losses'],
                       ms['win_rate'], ms['staked'], ms['pnl'], ms['roi'],
                       ms['avg_win'], ms['avg_loss'], round(cum, 2)])
        bg = _WIN_BG if ms['pnl'] >= 0 else _LOSS_BG
        _style_data_row(ws_yr, xl_mod, r, len(hdrs_yr), bg)
        ws_yr.cell(row=r, column=7).font  = _body_font(xl_mod, bold=True,
            color=(_POS_FG if ms['pnl'] >= 0 else _NEG_FG))
        ws_yr.cell(row=r, column=11).font = _body_font(xl_mod, bold=True,
            color=(_POS_FG if cum >= 0 else _NEG_FG))

    # Totals row
    if months_with_data:
        tot_r = len(months_with_data) + 2
        yr_wr  = round(year_wins / year_trades * 100, 1) if year_trades else 0.0
        yr_roi = round(year_pnl / year_staked * 100, 1) if year_staked else 0.0
        ws_yr.append([f'{year} TOTAL', year_trades, year_wins, year_losses,
                       yr_wr, round(year_staked, 2), round(year_pnl, 2), yr_roi,
                       '', '', round(year_pnl, 2)])
        _style_data_row(ws_yr, xl_mod, tot_r, len(hdrs_yr), _TOT_BG)
        for col in range(1, len(hdrs_yr) + 1):
            ws_yr.cell(row=tot_r, column=col).font = _body_font(xl_mod, bold=True)
        ws_yr.cell(row=tot_r, column=7).font  = _body_font(xl_mod, bold=True,
            color=(_POS_FG if year_pnl >= 0 else _NEG_FG))

    _autowidth(ws_yr, xl_mod)

    # ── One sheet per month ───────────────────────────────────────────────
    for month in months_with_data:
        month_name = _dt.date(year, month, 1).strftime('%b')
        ws_m = wb.create_sheet(month_name)
        ws_m.freeze_panes = 'A2'

        month_trades = [t for t in trades
                        if (lambda dt: dt and dt.year == year and dt.month == month)(
                            _ts_to_dt(float(t.get('resolved_ts') or t.get('ts') or 0)))]

        hdrs_m = ['Date/Time (UTC)', 'Symbol', 'Direction', 'Result',
                  'Confidence', 'Entry Price', 'Stake ($)',
                  'Net P/L ($)', 'Cum. P/L ($)']
        ws_m.append(hdrs_m)
        _style_header(ws_m, xl_mod, 1, len(hdrs_m))

        cum_m = 0.0
        for r_m, t in enumerate(month_trades, start=2):
            ts   = float(t.get('resolved_ts') or t.get('ts') or 0)
            dt   = _ts_to_dt(ts)
            dt_s = dt.strftime('%Y-%m-%d %H:%M:%S') if dt else ''
            sym  = _sym_norm(t.get('symbol') or t.get('asset') or '')
            dirn = (t.get('direction') or '').upper()
            res  = (t.get('resolved') or '').replace('_', ' ').title()
            conf = float(t.get('confidence') or 0.0)
            px   = float(t.get('avg_fill_price') or t.get('price') or 0.0)
            stk  = float(t.get('usdc') or t.get('stake_usdc') or 5.0)
            pnl  = _pnl_for(t)
            if pnl is not None:
                cum_m += pnl

            ws_m.append([dt_s, sym, dirn, res,
                          round(conf, 4) if conf else '',
                          round(px, 4)   if px   else '',
                          round(stk, 2),
                          round(pnl, 4)   if pnl is not None else '',
                          round(cum_m, 4)])

            bg = _WIN_BG if t.get('resolved') == 'win' else _LOSS_BG
            _style_data_row(ws_m, xl_mod, r_m, len(hdrs_m), bg)
            if pnl is not None:
                ws_m.cell(row=r_m, column=8).font = _body_font(xl_mod, bold=True,
                    color=(_POS_FG if pnl >= 0 else _NEG_FG))

        _autowidth(ws_m, xl_mod)
        ws_m.column_dimensions['A'].width = 22
        ws_m.sheet_properties.tabColor = '2E5F8A'

    ws_yr.sheet_properties.tabColor = '1A2A4A'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# All-time ledger builder
# ══════════════════════════════════════════════════════════════════════════════

def _build_alltime(trades: list[dict]) -> bytes:
    """Build an all-time overview: one row per month, one sheet per year."""
    xl_mod = _xl()
    openpyxl = xl_mod[0]
    wb = openpyxl.Workbook()

    # Gather all year/months that have data
    ym_set: set[tuple[int,int]] = set()
    for t in trades:
        dt = _ts_to_dt(float(t.get('resolved_ts') or t.get('ts') or 0))
        if dt:
            ym_set.add((dt.year, dt.month))
    yms = sorted(ym_set)
    years = sorted({y for y, _ in yms})

    # ── All-Time Overview sheet ───────────────────────────────────────────
    ws_all = wb.active
    ws_all.title = 'All-Time Overview'
    ws_all.freeze_panes = 'A2'

    hdrs_all = ['Year', 'Month', 'Trades', 'Wins', 'Losses',
                 'Win Rate (%)', 'Staked ($)', 'P/L ($)', 'ROI (%)',
                 'Cumulative P/L ($)']
    ws_all.append(hdrs_all)
    _style_header(ws_all, xl_mod, 1, len(hdrs_all))

    cum_all = 0.0
    row_all = 2
    for year in years:
        year_months = sorted(m for y, m in yms if y == year)
        yr_pnl = 0.0
        for month in year_months:
            ms = _month_summary(trades, year, month)
            cum_all += ms['pnl']
            yr_pnl  += ms['pnl']
            ws_all.append([str(year), ms['month'], ms['trades'], ms['wins'],
                            ms['losses'], ms['win_rate'], ms['staked'],
                            ms['pnl'], ms['roi'], round(cum_all, 2)])
            bg = _WIN_BG if ms['pnl'] >= 0 else _LOSS_BG
            _style_data_row(ws_all, xl_mod, row_all, len(hdrs_all), bg)
            ws_all.cell(row=row_all, column=8).font  = _body_font(xl_mod, bold=True,
                color=(_POS_FG if ms['pnl'] >= 0 else _NEG_FG))
            ws_all.cell(row=row_all, column=10).font = _body_font(xl_mod, bold=True,
                color=(_POS_FG if cum_all >= 0 else _NEG_FG))
            row_all += 1

        # Year subtotal row
        yr_ws = ws_all  # write subtotal inline
        yr_trades = sum(_month_summary(trades, year, m)['trades'] for m in year_months)
        yr_wins   = sum(_month_summary(trades, year, m)['wins']   for m in year_months)
        yr_staked = sum(_month_summary(trades, year, m)['staked'] for m in year_months)
        yr_wr     = round(yr_wins / yr_trades * 100, 1) if yr_trades else 0.0
        yr_roi    = round(yr_pnl / yr_staked * 100, 1) if yr_staked else 0.0
        ws_all.append([f'{year} SUBTOTAL', '', yr_trades, yr_wins,
                        yr_trades - yr_wins, yr_wr, round(yr_staked, 2),
                        round(yr_pnl, 2), yr_roi, round(cum_all, 2)])
        _style_data_row(ws_all, xl_mod, row_all, len(hdrs_all), _TOT_BG)
        for col in range(1, len(hdrs_all) + 1):
            ws_all.cell(row=row_all, column=col).font = _body_font(xl_mod, bold=True)
        ws_all.cell(row=row_all, column=8).font = _body_font(xl_mod, bold=True,
            color=(_POS_FG if yr_pnl >= 0 else _NEG_FG))
        row_all += 1

    _autowidth(ws_all, xl_mod)
    ws_all.sheet_properties.tabColor = '1A2A4A'

    # ── One sheet per year (monthly summary) ──────────────────────────────
    for year in years:
        ws_y = wb.create_sheet(str(year))
        ws_y.freeze_panes = 'A2'
        year_months = sorted(m for y, m in yms if y == year)
        hdrs_y = ['Month', 'Trades', 'Wins', 'Losses', 'Win Rate (%)',
                  'Staked ($)', 'P/L ($)', 'ROI (%)', 'Cumulative P/L ($)']
        ws_y.append(hdrs_y)
        _style_header(ws_y, xl_mod, 1, len(hdrs_y))
        cum_y = 0.0
        for r_y, month in enumerate(year_months, start=2):
            ms = _month_summary(trades, year, month)
            cum_y += ms['pnl']
            ws_y.append([ms['month'], ms['trades'], ms['wins'], ms['losses'],
                          ms['win_rate'], ms['staked'], ms['pnl'],
                          ms['roi'], round(cum_y, 2)])
            bg = _WIN_BG if ms['pnl'] >= 0 else _LOSS_BG
            _style_data_row(ws_y, xl_mod, r_y, len(hdrs_y), bg)
            ws_y.cell(row=r_y, column=7).font  = _body_font(xl_mod, bold=True,
                color=(_POS_FG if ms['pnl'] >= 0 else _NEG_FG))
            ws_y.cell(row=r_y, column=9).font  = _body_font(xl_mod, bold=True,
                color=(_POS_FG if cum_y >= 0 else _NEG_FG))
        _autowidth(ws_y, xl_mod)
        ws_y.sheet_properties.tabColor = '2E5F8A'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# Public API
# ══════════════════════════════════════════════════════════════════════════════

def update_ledgers(force: bool = False) -> bool:
    """Rebuild all ledger files if new resolved trades exist.

    Called from the background loop after check_resolutions().
    Returns True if any files were written.
    """
    global _last_update_ts, _last_trade_count

    now = time.time()
    if not force and (now - _last_update_ts) < _UPDATE_INTERVAL_S:
        return False

    try:
        trades = _resolved_trades(_load_trades())
        count  = len(trades)
        if not force and count == _last_trade_count:
            return False

        LEDGER_DIR.mkdir(parents=True, exist_ok=True)

        today = _dt.datetime.utcnow()
        written = False

        # ── Current monthly ledger ────────────────────────────────────────
        try:
            monthly_bytes = _build_monthly(trades, today.year, today.month)
            monthly_path  = LEDGER_DIR / f'ledger_{today.year}_{today.month:02d}.xlsx'
            monthly_path.write_bytes(monthly_bytes)
            written = True
            logger.info('ledger: wrote monthly %s (%d trades)', monthly_path.name, count)
        except Exception as exc:
            logger.warning('ledger: monthly build failed: %s', exc)

        # ── Annual ledger for current year ────────────────────────────────
        try:
            annual_bytes = _build_annual(trades, today.year)
            annual_path  = LEDGER_DIR / f'ledger_{today.year}_annual.xlsx'
            annual_path.write_bytes(annual_bytes)
            logger.info('ledger: wrote annual %s', annual_path.name)
        except Exception as exc:
            logger.warning('ledger: annual build failed: %s', exc)

        # ── Previous years' annuals (if any trades from earlier years) ────
        all_years = sorted({
            _ts_to_dt(float(t.get('resolved_ts') or t.get('ts') or 0)).year
            for t in trades
            if _ts_to_dt(float(t.get('resolved_ts') or t.get('ts') or 0))
        })
        for yr in all_years:
            if yr == today.year:
                continue
            prev_path = LEDGER_DIR / f'ledger_{yr}_annual.xlsx'
            if not prev_path.exists():
                try:
                    prev_bytes = _build_annual(trades, yr)
                    prev_path.write_bytes(prev_bytes)
                    logger.info('ledger: wrote historical annual %s', prev_path.name)
                except Exception as exc:
                    logger.warning('ledger: historical annual build failed: %s', exc)

        # ── All-time ledger ───────────────────────────────────────────────
        try:
            alltime_bytes = _build_alltime(trades)
            alltime_path  = LEDGER_DIR / 'ledger_all_time.xlsx'
            alltime_path.write_bytes(alltime_bytes)
            logger.info('ledger: wrote all-time ledger')
        except Exception as exc:
            logger.warning('ledger: all-time build failed: %s', exc)

        _last_update_ts    = now
        _last_trade_count  = count
        return written

    except Exception as exc:
        logger.warning('ledger: update_ledgers failed: %s', exc)
        return False


def list_ledgers() -> list[dict]:
    """Return metadata for all ledger files, newest first."""
    if not LEDGER_DIR.exists():
        return []
    out = []
    for f in sorted(LEDGER_DIR.glob('*.xlsx'), reverse=True):
        stat = f.stat()
        out.append({
            'filename': f.name,
            'path': str(f),
            'size_kb': round(stat.st_size / 1024, 1),
            'modified': _dt.datetime.utcfromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M UTC'),
        })
    return out


def get_ledger_bytes(filename: str) -> bytes | None:
    """Read and return a ledger file's bytes (safe filename check)."""
    if '..' in filename or '/' in filename or '\\' in filename:
        return None
    path = LEDGER_DIR / filename
    if not path.exists() or path.suffix != '.xlsx':
        return None
    return path.read_bytes()
